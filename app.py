import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import os

# Set page configurations and metadata for SEO and browser display
st.set_page_config(
    page_title="Private Sports Lessons Workbook Generator",
    page_icon="🎾",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Initialize session state variables
if "generated" not in st.session_state:
    st.session_state.generated = False

def reset_generation():
    st.session_state.generated = False

# Premium theme injection using modern typography, radial gradients, and custom button micro-animations
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;600;800&display=swap');
html, body, [data-testid="stAppViewContainer"] {
    font-family: 'Outfit', sans-serif;
    background: radial-gradient(circle at 10% 20%, rgba(26, 42, 62, 1) 0%, rgba(15, 23, 36, 1) 90.1%);
    color: #f8fafc;
}
[data-testid="stSidebar"] {
    background-color: rgba(15, 23, 36, 0.8) !important;
    backdrop-filter: blur(10px);
    border-right: 1px solid rgba(255, 255, 255, 0.05);
}
h1, h2, h3 {
    color: #38bdf8 !important;
    font-weight: 800 !important;
}
.stRadio > div {
    flex-direction: row;
}
div.stButton > button {
    background: linear-gradient(135deg, #0284c7 0%, #0369a1 100%) !important;
    color: white !important;
    border: none !important;
    padding: 12px 28px !important;
    font-size: 16px !important;
    font-weight: 600 !important;
    border-radius: 8px !important;
    transition: all 0.3s ease !important;
    box-shadow: 0 4px 12px rgba(2, 132, 199, 0.3) !important;
    width: 100%;
}
div.stButton > button:hover {
    transform: translateY(-2px) !important;
    box-shadow: 0 6px 20px rgba(2, 132, 199, 0.5) !important;
}
.card {
    background-color: rgba(255, 255, 255, 0.03);
    border: 1px solid rgba(255, 255, 255, 0.05);
    border-radius: 12px;
    padding: 24px;
    margin-bottom: 20px;
}
.metric-box {
    background-color: rgba(56, 189, 248, 0.08);
    border: 1px solid rgba(56, 189, 248, 0.2);
    border-radius: 8px;
    padding: 12px;
    text-align: center;
}
</style>
""", unsafe_allow_html=True)

# Default Counselor Mapping extracted from the updated Session 2 sheet
DEFAULT_COUNSELORS = {}

def _add_counselor(bunks, counselor):
    for b in bunks:
        DEFAULT_COUNSELORS[b] = counselor

# Katie McCall: GJL1 - GJL3
_add_counselor([f"GJL{i}" for i in range(1, 4)], "Katie McCall")
# Sophie Boswell: GJL4 - GJL6
_add_counselor([f"GJL{i}" for i in range(4, 7)], "Sophie Boswell")
# Harlee Walker: G1 & G2
_add_counselor([f"G{i:02d}{s}" for i in range(1, 3) for s in ['A','B','C']], "Harlee Walker")
# Jessica Morrison: G3 & G4
_add_counselor([f"G{i:02d}{s}" for i in range(3, 5) for s in ['A','B','C']], "Jessica Morrison")
# Kirsten Gouldie: G5 & G6
_add_counselor([f"G{i:02d}{s}" for i in range(5, 7) for s in ['A','B','C']], "Kirsten Gouldie")
# Alyssa Shanks: G7 & G8
_add_counselor([f"G{i:02d}{s}" for i in range(7, 9) for s in ['A','B','C']], "Alyssa Shanks")
# Liz Hernandez: G9 & G10
_add_counselor([f"G{i:02d}{s}" for i in range(9, 11) for s in ['A','B','C']], "Liz Hernandez")
# Cara McMahon: G11 - G14
_add_counselor([f"G{i:02d}{s}" for i in range(11, 15) for s in ['A','B','C']], "Cara McMahon")
# Savannah Elphick: G15 & G16
_add_counselor([f"G{i:02d}{s}" for i in range(15, 17) for s in ['A','B','C']], "Savannah Elphick")
# Hannah Roche: G17 - G19
_add_counselor([f"G{i:02d}{s}" for i in range(17, 20) for s in ['A','B','C']], "Hannah Roche")
# Claudia Anderson: GSL1 - 4
_add_counselor([f"GSL{i}" for i in range(1, 5)], "Claudia Anderson")

# Nathaniel Eldridge-Smith: BJL1 - B2
_add_counselor([f"BJL{i}" for i in range(1, 4)], "Nathaniel Eldridge-Smith")
_add_counselor([f"B{i:02d}{s}" for i in range(1, 3) for s in ['A','B','C']], "Nathaniel Eldridge-Smith")
# Justyn Allen: B3 - B5
_add_counselor([f"B{i:02d}{s}" for i in range(3, 6) for s in ['A','B','C']], "Justyn Allen")
# Jake Seymore: B6 & B7
_add_counselor([f"B{i:02d}{s}" for i in range(6, 8) for s in ['A','B','C']], "Jake Seymore")
# Marc Marsh: B8 & B9
_add_counselor([f"B{i:02d}{s}" for i in range(8, 10) for s in ['A','B','C']], "Marc Marsh")
# Josh Wilson: B10 - BSL2
_add_counselor([f"B{i:02d}{s}" for i in range(10, 16) for s in ['A','B','C']], "Josh Wilson")
_add_counselor([f"BSL{i}" for i in range(1, 3)], "Josh Wilson")

# Fixed Calendar Configurations for each session
DATES_BY_SESSION = {
    "Session 1": [
        "9-Jun", "10-Jun", "11-Jun", "12-Jun", "13-Jun", "15-Jun", "16-Jun", 
        "17-Jun", "18-Jun", "19-Jun", "20-Jun", "21-Jun", "22-Jun", "23-Jun"
    ],
    "Session 2": [
        '27-Jun', '28 -Jun', '29 -Jun', '1 -July', '2 -July', '3 -July', 
        '4 -July', '5 -July', '7 -July', '8 -July', '9 -July', '10 -July', 
        '11 -July', '12 -July'
    ],
    "Session 3": [
        # --- PLACEHOLDER: Easily customize dates for Session 3 below ---
        "13-July", "14-July", "15-July", "16-July", "17-July", "18-July",
        "20-July", "21-July", "22-July", "23-July", "24-July", "25-July"
    ],
    "Session 4": [
        # --- PLACEHOLDER: Easily customize dates for Session 4 below ---
        "27-July", "28-July", "29-July", "30-July", "31-July", "1-Aug",
        "3-Aug", "4-Aug", "5-Aug", "6-Aug", "7-Aug", "8-Aug"
    ]
}

def clean_session_name(sess_name):
    """Convert long enrollment session descriptions into compact shorthands."""
    if pd.isnull(sess_name):
        return ""
    sess_str = str(sess_name).strip()
    mapping = {
        "Session 1": "S1",
        "Session 2": "S2",
        "Session 3": "S3",
        "Session 4": "S4",
        "Sessions 1 & 2": "S1&2",
        "Sessions 2 & 3": "S2&3",
        "Sessions 3 & 4": "S3&4",
        "Sessions 2  3 & 4": "Sessions 2  3 & 4",
        "Full Summer (1-4)": "Full Summer"
    }
    return mapping.get(sess_str, sess_str)

def process_roster(session_df, sport, session_name, counselor_map):
    """
    Roster Filtering Logic:
    1. Filter out campers with 0 or NaN session lessons or cancellation memos.
    2. Sort alphabetically by Last Name, then First Name.
    3. Perform 'oder session hours' math and return rows with unified values.
    """
    # Sport configurations
    sport_prefix = "T" if sport == "Tennis" else "P"
    
    # Base column lookups
    level_col_name = f"{sport}Level"
    total_lessons_col_name = f"{sport}TotalLessons"
    memo_col_name = f"{sport_prefix}-{sport}Memo"
    
    session_num_str = "".join([c for c in session_name if c.isdigit()])
    sess_lessons_col = f"{sport}Session{session_num_str}Lessons"
    
    def find_column_case_insensitive(df, target_name):
        """Find df column matching target_name case-insensitively, falling back to substring search."""
        for col in df.columns:
            if col.strip().lower() == target_name.lower():
                return col
        for col in df.columns:
            if target_name.lower() in col.lower():
                return col
        return None

    # Resolve session column names
    bunk_col = None
    for col in session_df.columns:
        if "bunk" in col.lower():
            bunk_col = col
            break
    if not bunk_col:
        bunk_col = session_df.columns[0]
        
    sess_lessons_col_actual = find_column_case_insensitive(session_df, sess_lessons_col)
    if not sess_lessons_col_actual:
        # Fallback substring matching for current session lessons
        for col in session_df.columns:
            if session_num_str in col and sport.lower() in col.lower() and "lessons" in col.lower():
                sess_lessons_col_actual = col
                break
    if not sess_lessons_col_actual:
        sess_lessons_col_actual = sess_lessons_col
        
    sess_total_col_actual = find_column_case_insensitive(session_df, total_lessons_col_name) or find_column_case_insensitive(session_df, "total") or total_lessons_col_name
    sess_level_col_actual = find_column_case_insensitive(session_df, level_col_name) or find_column_case_insensitive(session_df, "level") or level_col_name
    sess_memo_col_actual = find_column_case_insensitive(session_df, memo_col_name) or find_column_case_insensitive(session_df, "memo")
    
    # Parse lessons for current session
    session_df['current_sess_lessons'] = pd.to_numeric(session_df[sess_lessons_col_actual], errors='coerce').fillna(0)
    
    # Identify cancelled lessons
    if sess_memo_col_actual and sess_memo_col_actual in session_df.columns:
        memos = session_df[sess_memo_col_actual].fillna('').astype(str)
    else:
        memos = pd.Series([''] * len(session_df))
    is_cancelled = memos.str.contains(r'cancel|don\'t charge|no charge', case=False, na=False)
    
    # Filter active camper list
    active = session_df[(session_df['current_sess_lessons'] > 0) & (~is_cancelled)].copy()
    
    # Sort alphabetically by last name then first name (case-insensitive)
    last_name_col = find_column_case_insensitive(active, 'Last Name') or 'Last Name'
    first_name_col = find_column_case_insensitive(active, 'First Name') or 'First Name'
    if last_name_col in active.columns and first_name_col in active.columns:
        active['_sort_last'] = active[last_name_col].astype(str).str.lower()
        active['_sort_first'] = active[first_name_col].astype(str).str.lower()
        active = active.sort_values(by=['_sort_last', '_sort_first']).drop(columns=['_sort_last', '_sort_first']).reset_index(drop=True)
    
    # Build uniform camper roster structure
    roster_rows = []
    for idx, row in active.iterrows():
        bunk = str(row[bunk_col]).strip() if pd.notnull(row.get(bunk_col)) else ""
        
        # Look up counselor mapping
        counselor = counselor_map.get((session_name, bunk), counselor_map.get(bunk, ""))
        
        last_name = row.get(last_name_col, "")
        first_name = row.get(first_name_col, "")
        
        # Age
        age = row.get('Age', "")
        if pd.notnull(age) and age != "":
            try:
                age = float(age)
            except ValueError:
                pass
        else:
            age = ""
            
        # Session shorthand conversion
        raw_sess = row.get('Enrolled Sessions', "")
        clean_sess = clean_session_name(raw_sess)
        
        # Skill level
        level = row.get(sess_level_col_actual, "") if sess_level_col_actual else ""
        if pd.isnull(level):
            level = ""
        level = str(level).strip()
        
        # Minor Availability (empty placeholder column by default)
        minor_availability = ""
        
        # Memo
        memo = row.get(sess_memo_col_actual, "") if sess_memo_col_actual else ""
        if pd.isnull(memo):
            memo = ""
        memo = str(memo).strip()
        
        session_lessons = row.get('current_sess_lessons', 0)
        
        # Total lessons resolution
        total_lessons = row.get(sess_total_col_actual) if sess_total_col_actual else None
        if pd.isnull(total_lessons) or total_lessons == 0 or total_lessons == "":
            total_lessons = session_lessons
            
        try:
            total_lessons = float(total_lessons)
        except (ValueError, TypeError):
            total_lessons = session_lessons
                
        roster_rows.append({
            "bunk": bunk,
            "counselor": counselor,
            "last_name": last_name,
            "first_name": first_name,
            "age": age,
            "session": clean_sess,
            "level": level,
            "minor_availability": minor_availability,
            "memo": memo,
            "session_lessons": int(session_lessons),
            "total_lessons": int(total_lessons)
        })
        
    return roster_rows

def generate_excel_workbook(roster_rows, sport, session_name):
    """
    Generate the Excel file using openpyxl, applying formatting styles, widths,
    formulas, and bottom summary fields to perfectly match the design specifications.
    """
    wb = openpyxl.Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Create the session worksheet
    ws = wb.create_sheet(title=session_name)
    ws.views.sheetView[0].showGridLines = True  # Ensure gridlines are visible
    
    # Build column configurations
    session_num_str = "".join([c for c in session_name if c.isdigit()])
    sport_prefix = "T" if sport == "Tennis" else "P"
    
    memo_header = f"{sport_prefix}-{sport}Memo"
    session_total_header = f"Session {session_num_str} {sport} Lessons TOTAL"
    
    headers = [
        "#", "Bunk", "Head Counselor", "Last Name", "First Name", "Age",
        "session", "Skill Level", "Minor Availability", memo_header,
        session_total_header, "oder session hours"
    ]
    
    dates = DATES_BY_SESSION.get(session_name, [])
    headers.extend(dates)
    
    trailing_headers = [
        "Programming Notes", "Refund Amount", "Refund Check # or Clover"
    ]
    headers.extend(trailing_headers)
    
    # 1. Write and Style Title Cell (Row 1)
    title_text = f"PRIVATE {sport.upper()} LESSONS \n{session_name.upper()} 2026"
    ws.cell(row=1, column=1, value=title_text)
    
    last_col_idx = len(headers)
    last_col_letter = get_column_letter(last_col_idx)
    ws.merge_cells(f"A1:{last_col_letter}1")
    
    title_cell = ws["A1"]
    title_cell.font = Font(name="Arial", size=36.0, bold=True, color="000000")
    title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 96.0
    
    # 2. Write and Style Header Row (Row 2)
    ws.row_dimensions[2].height = 30.0
    header_fill = PatternFill(start_color="AEAAAA", end_color="AEAAAA", fill_type="solid")
    header_font_small = Font(name="Arial", size=10.0, bold=True, color="000000")
    header_font_large = Font(name="Arial", size=14.0, bold=True, color="000000")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    for col_idx, h_text in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=h_text)
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        
        # Bigger font for dates in columns
        if 13 <= col_idx <= 12 + len(dates):
            cell.font = header_font_large
        else:
            cell.font = header_font_small
            
    # 3. Write and Style Camper Rows (Row 3 onwards)
    fill_k = PatternFill(start_color="FFD0D4D9", end_color="FFD0D4D9", fill_type="solid")
    fill_l = PatternFill(start_color="FFE6EBF0", end_color="FFE6EBF0", fill_type="solid")
    font_l = Font(name="Arial", size=18.0, bold=True, color="FF7B8187")
    
    row_idx = 3
    for row_data in roster_rows:
        ws.row_dimensions[row_idx].height = 70.5
        
        # Sequence Number
        cell_num = ws.cell(row=row_idx, column=1, value=row_idx - 2)
        cell_num.font = Font(name="Arial", size=14.0, bold=True, color="000000")
        cell_num.alignment = Alignment(horizontal="center", vertical="center")
        
        # Bunk
        cell_bunk = ws.cell(row=row_idx, column=2, value=row_data["bunk"])
        cell_bunk.font = Font(name="Arial", size=14.0, bold=False, color="000000")
        cell_bunk.alignment = Alignment(horizontal="center", vertical="center")
        
        # Head Counselor
        cell_counselor = ws.cell(row=row_idx, column=3, value=row_data["counselor"])
        cell_counselor.font = Font(name="Arial", size=12.0, bold=False, color="000000")
        cell_counselor.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Last Name
        cell_ln = ws.cell(row=row_idx, column=4, value=row_data["last_name"])
        cell_ln.font = Font(name="Arial", size=14.0, bold=False, color="000000")
        cell_ln.alignment = Alignment(horizontal="center", vertical="center")
        
        # First Name
        cell_fn = ws.cell(row=row_idx, column=5, value=row_data["first_name"])
        cell_fn.font = Font(name="Arial", size=14.0, bold=False, color="000000")
        cell_fn.alignment = Alignment(horizontal="center", vertical="center")
        
        # Age
        cell_age = ws.cell(row=row_idx, column=6, value=row_data["age"])
        cell_age.font = Font(name="Arial", size=14.0, bold=False, color="000000")
        cell_age.alignment = Alignment(horizontal="center", vertical="center")
        if isinstance(row_data["age"], (int, float)):
            cell_age.number_format = '0.00'
            
        # Session shorthand
        cell_sess = ws.cell(row=row_idx, column=7, value=row_data["session"])
        cell_sess.font = Font(name="Arial", size=16.0, bold=False, color="000000")
        cell_sess.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # Skill level
        cell_lvl = ws.cell(row=row_idx, column=8, value=row_data["level"])
        cell_lvl.font = Font(name="Arial", size=12.0, bold=False, color="000000")
        cell_lvl.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # Minor availability
        cell_avail = ws.cell(row=row_idx, column=9, value=row_data["minor_availability"])
        cell_avail.font = Font(name="Arial", size=14.0, bold=True)
        cell_avail.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Sport Memo
        cell_memo = ws.cell(row=row_idx, column=10, value=row_data["memo"])
        cell_memo.font = Font(name="Arial", size=10.0, bold=False, color="000000")
        cell_memo.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        session_lessons_val = row_data["session_lessons"]
        date_start_letter = get_column_letter(13)
        date_end_letter = get_column_letter(12 + len(dates))
        
        # Session total formula (Col 11 / K) - Total minus Sum of dates
        sess_total_formula = f"={session_lessons_val}-SUM({date_start_letter}{row_idx}:{date_end_letter}{row_idx})"
        cell_tot = ws.cell(row=row_idx, column=11, value=sess_total_formula)
        cell_tot.font = Font(name="Arial", size=24.0, bold=False, color="000000")
        cell_tot.alignment = Alignment(horizontal="center", vertical="center")
        cell_tot.fill = fill_k
        
        # Other session hours formula (Col 12 / L) - Total minus the starting session lessons
        other_hours_formula = f"={row_data['total_lessons']}-{session_lessons_val}"
        cell_other = ws.cell(row=row_idx, column=12, value=other_hours_formula)
        cell_other.font = font_l
        cell_other.alignment = Alignment(horizontal="center", vertical="center")
        cell_other.fill = fill_l
        
        # Dates (Cols 13 to 12 + len(dates)) - empty grid spaces
        for c in range(13, 13 + len(dates)):
            cell_d = ws.cell(row=row_idx, column=c)
            cell_d.font = Font(name="Arial", size=14.0, bold=False)
            cell_d.alignment = Alignment(horizontal="center", vertical="center")
            cell_d.border = thin_border
            
        # Trailing columns: Notes, Refund, Refund Clover (starts directly after dates)
        for c in range(13 + len(dates), 16 + len(dates)):
            cell_trail = ws.cell(row=row_idx, column=c)
            cell_trail.font = Font(name="Arial", size=14.0, bold=False)
            cell_trail.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
        # Apply borders and white fills to all columns in the row
        white_fill = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.border = thin_border
            if c == 1:
                cell.fill = white_fill
            
        row_idx += 1
        
    last_camper_row = row_idx - 1
    
    # 4. Totals Sum Row (Row last_camper_row + 2)
    totals_row = last_camper_row + 2
    ws.row_dimensions[totals_row].height = 20.0
    
    cell_totals_label = ws.cell(row=totals_row, column=1, value="TOTALS")
    cell_totals_label.font = Font(name="Arial", size=16.0, bold=True, color="000000")
    cell_totals_label.alignment = Alignment(horizontal="center", vertical="center")
    
    cell_totals_k = ws.cell(row=totals_row, column=11, value=f"=SUM(K3:K{last_camper_row})")
    cell_totals_k.font = Font(name="Arial", size=16.0, bold=True, color="000000")
    cell_totals_k.alignment = Alignment(horizontal="center", vertical="center")
    cell_totals_k.fill = fill_k
    
    cell_totals_l = ws.cell(row=totals_row, column=12, value=f"=SUM(L3:L{last_camper_row})")
    cell_totals_l.font = Font(name="Arial", size=18.0, bold=True, color="7B8187")
    cell_totals_l.alignment = Alignment(horizontal="center", vertical="center")
    cell_totals_l.fill = fill_l
    
    for c in range(1, len(headers) + 1):
        ws.cell(row=totals_row, column=c).border = thin_border
        
    # 5. Show-up and Cancellation Tracker Boxes (below table)
    showup_row = last_camper_row + 4
    cancel_row = last_camper_row + 5
    
    ws.merge_cells(start_row=showup_row, start_column=9, end_row=showup_row, end_column=10)
    cell_showup_lbl = ws.cell(row=showup_row, column=9, value="THE CAMPER DIDN'T SHOW UP FOR THE LESSON TOTAL:")
    cell_showup_lbl.font = Font(name="Arial", size=12.0, bold=True, color="000000")
    cell_showup_lbl.alignment = Alignment(horizontal="right", vertical="center")
    
    ws.merge_cells(start_row=cancel_row, start_column=9, end_row=cancel_row, end_column=10)
    cell_cancel_lbl = ws.cell(row=cancel_row, column=9, value="LESSONS CANCELED DUE TO RAIN TOTAL:")
    cell_cancel_lbl.font = Font(name="Arial", size=12.0, bold=True, color="000000")
    cell_cancel_lbl.alignment = Alignment(horizontal="right", vertical="center")
    
    ws.merge_cells(start_row=showup_row, start_column=11, end_row=showup_row, end_column=13)
    ws.merge_cells(start_row=cancel_row, start_column=11, end_row=cancel_row, end_column=13)
    
    for r in [showup_row, cancel_row]:
        # Style label merges borders
        ws.cell(row=r, column=9).border = Border(left=Side(style='thin', color='000000'), top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000'))
        ws.cell(row=r, column=10).border = Border(right=Side(style='thin', color='000000'), top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000'))
        
        # Style box merges borders
        ws.cell(row=r, column=11).border = Border(left=Side(style='thin', color='000000'), top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000'))
        ws.cell(row=r, column=12).border = Border(top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000'))
        ws.cell(row=r, column=13).border = Border(right=Side(style='thin', color='000000'), top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000'))
        
    # 6. Apply Exact Column Widths
    column_widths = {
        "A": 8, "B": 10, "C": 15, "D": 18, "E": 18, "F": 10, "G": 12, "H": 15, "I": 15, "J": 35, "K": 12, "L": 12
    }
    for c in range(13, 13 + len(dates)):
        col_letter = get_column_letter(c)
        column_widths[col_letter] = 10
    
    trail_letters = [get_column_letter(c) for c in range(13 + len(dates), 16 + len(dates))]
    trail_widths = [20, 12, 18]  # Notes, Refund, Check info
    for letter, w in zip(trail_letters, trail_widths):
        column_widths[letter] = w
        
    for letter, width in column_widths.items():
        ws.column_dimensions[letter].width = width
        
    # Return workbook saved in-memory
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return file_stream


# --- STREAMLIT UI LAYOUT ---

# App Header Card
st.markdown("""
<div class="card">
    <h1 style="margin:0; font-size: 32px; letter-spacing: -0.5px;">🎾 Private Sports Lessons Workbook Generator</h1>
    <p style="margin:5px 0 0 0; color: #94a3b8; font-size: 16px;">
        Automate the generation of private sports lessons scheduling workbooks for Tennis and Pickleball registrations.
    </p>
</div>
""", unsafe_allow_html=True)

# Main Grid layout: Sidebar + Data panel
st.sidebar.markdown("## ⚙️ Configuration")

# Sport Selection
sport = st.sidebar.radio(
    "Select Sport:",
    ["Tennis", "Pickleball"],
    index=0,
    help="Underlying logic is identical; text labels, column names, and sheet titles adjust dynamically.",
    key="sport_select",
    on_change=reset_generation
)

# Session Selection
session_name = st.sidebar.selectbox(
    "Select Session:",
    ["Session 1", "Session 2", "Session 3", "Session 4"],
    index=1,  # Default to Session 2
    help="Determines the calendar dates appended to the roster sheet.",
    key="session_select",
    on_change=reset_generation
)

# Counselor Mapping Section in Sidebar (Interactive Data Editor)
st.sidebar.markdown("### 👤 Head Counselors")
st.sidebar.write("Customize head counselors for each bunk. Add or modify entries as needed:")

# Initialize session state for counselor mappings if not present
if "counselor_data_v2" not in st.session_state:
    flat_list = []
    for k, v in DEFAULT_COUNSELORS.items():
        if isinstance(k, tuple):
            flat_list.append({"Session": k[0], "Bunk": k[1], "Head Counselor": v})
        else:
            flat_list.append({"Session": "Global Fallback", "Bunk": k, "Head Counselor": v})
    st.session_state.counselor_data_v2 = pd.DataFrame(flat_list)

edited_df = st.sidebar.data_editor(
    st.session_state.counselor_data_v2,
    num_rows="dynamic",
    use_container_width=True,
    key="counselor_editor_widget_v2",
    on_change=reset_generation
)

# Convert edited data frame back to dict mapping
counselor_map = {}
for _, row in edited_df.iterrows():
    sess = row.get("Session")
    bunk = str(row.get("Bunk")).strip() if pd.notnull(row.get("Bunk")) else ""
    counselor = str(row.get("Head Counselor")).strip() if pd.notnull(row.get("Head Counselor")) else ""
    if bunk:
        if sess and sess != "Global Fallback":
            counselor_map[(sess, bunk)] = counselor
        else:
            counselor_map[bunk] = counselor

# File Upload Section
st.markdown("### 📂 Data Ingestion")
st.markdown('<div class="card">', unsafe_allow_html=True)
st.markdown(f"#### 📅 Session specific export")
session_file = st.file_uploader(
    f"Upload session export CSV (e.g., session2tennis.csv)",
    type=["csv"],
    key="session_uploader",
    help="Contains bunk, age, total lessons, and lessons allocated for the selected session.",
    on_change=reset_generation
)
st.markdown('</div>', unsafe_allow_html=True)

# Local files fallback logic for development / quick testing
session_num_str = "".join([c for c in session_name if c.isdigit()])
if sport == "Tennis":
    local_session_name = f"session{session_num_str}tennis.csv"
else:
    local_session_name = f"pickleballSession{session_num_str}.csv"

local_session_exists = os.path.exists(local_session_name)

if local_session_exists:
    st.info(f"💡 Local database file {local_session_name} detected in workspace.")
    use_local = st.checkbox(
        "🔄 Automatically load local workspace CSV file for processing",
        value=True,
        key="use_local_checkbox",
        on_change=reset_generation
    )
else:
    use_local = False

# Processing Trigger
session_df = None

if use_local:
    try:
        session_df = pd.read_csv(local_session_name)
    except Exception as e:
        st.error(f"Failed to read local files: {e}")
else:
    if session_file is not None:
        try:
            session_df = pd.read_csv(session_file)
        except Exception as e:
            st.error(f"Error parsing Session CSV: {e}")

if session_df is not None:
    # Action button to generate workbook
    st.markdown("### ⚡ Actions")
    col_gen, _ = st.columns([1, 2])
    with col_gen:
        if st.button("🚀 Generate Workbook", key="generate_btn", type="primary", use_container_width=True):
            st.session_state.generated = True

    if st.session_state.generated:
        # Process Roster
        try:
            roster_rows = process_roster(session_df, sport, session_name, counselor_map)
            
            # Display Stats Summary Card
            st.markdown("### 📊 Roster Summary")
            col_m1, col_m2, col_m3 = st.columns(3)
            with col_m1:
                st.markdown(f'<div class="metric-box"><h3>Active Roster Size</h3><h2>{len(roster_rows)}</h2></div>', unsafe_allow_html=True)
            with col_m2:
                sess_tot = sum(r['session_lessons'] for r in roster_rows)
                st.markdown(f'<div class="metric-box"><h3>{sport} Session Lessons</h3><h2>{sess_tot}</h2></div>', unsafe_allow_html=True)
            with col_m3:
                other_tot = sum(r['total_lessons'] - r['session_lessons'] for r in roster_rows)
                st.markdown(f'<div class="metric-box"><h3>Other Sessions Lessons</h3><h2>{other_tot}</h2></div>', unsafe_allow_html=True)
                
            # Display Data Roster Preview Tab
            st.markdown("### 👁️ Processed Roster Preview")
            preview_df = pd.DataFrame(roster_rows)
            # Rename columns to match target Excel layout headers for preview
            preview_cols_rename = {
                "bunk": "Bunk",
                "counselor": "Head Counselor",
                "last_name": "Last Name",
                "first_name": "First Name",
                "age": "Age",
                "session": "session",
                "level": "Skill Level",
                "minor_availability": "Minor Availability",
                "memo": f"{'T' if sport == 'Tennis' else 'P'}-{sport}Memo",
                "session_lessons": f"Session {''.join([c for c in session_name if c.isdigit()])} {sport} Lessons TOTAL",
                "total_lessons": "Total Lessons"
            }
            
            preview_display = preview_df.rename(columns=preview_cols_rename)
            # Calculate other session hours for preview display
            preview_display["oder session hours"] = preview_display["Total Lessons"] - preview_display[preview_cols_rename["session_lessons"]]
            # Reorder to match visual layout
            ordered_preview_cols = [
                "Bunk", "Head Counselor", "Last Name", "First Name", "Age", "session", "Skill Level", 
                "Minor Availability", preview_cols_rename["memo"], preview_cols_rename["session_lessons"], 
                "oder session hours"
            ]
            
            st.dataframe(preview_display[ordered_preview_cols], use_container_width=True)
            
            # Generation and Download button
            excel_buffer = generate_excel_workbook(roster_rows, sport, session_name)
            
            filename = f"Private {sport} Lessons {session_name}.xlsx"
            
            st.markdown("<br>", unsafe_allow_html=True)
            col_btn, col_spacer = st.columns([1, 2])
            with col_btn:
                st.download_button(
                    label=f"📥 Download Formatted Excel Workbook",
                    data=excel_buffer,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                st.success(f"Workbook successfully prepared as '{filename}'!")
                
        except Exception as e:
            st.error(f"An error occurred during data processing: {e}")
            st.exception(e)
    else:
        st.info("👈 Click the **Generate Workbook** button above to process the data and compile the workbook.")
else:
    # Guidelines Card when files are missing
    st.markdown("""
    <div class="card" style="border-left: 5px solid #38bdf8;">
        <h3>📌 Getting Started</h3>
        <p>Please upload the Session specific CSV, or ensure local csv files are present in the workspace directory to generate the lessons schedule.</p>
        <ul>
            <li><strong>Data Cleansing:</strong> Excludes non-active and cancelled registrations dynamically.</li>
            <li><strong>Styling Replication:</strong> Applies custom colors, cell fonts, row heights, and layout borders.</li>
        </ul>
    </div>
    """, unsafe_allow_html=True)
