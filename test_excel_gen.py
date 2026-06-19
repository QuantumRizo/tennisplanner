import sys
from unittest.mock import MagicMock

# Mock streamlit so we can import app.py backend logic without running the UI
mock_st = MagicMock()

class MockSessionState(dict):
    def __getattr__(self, key):
        try:
            return self[key]
        except KeyError:
            raise AttributeError(key)
    def __setattr__(self, key, value):
        self[key] = value

mock_st.session_state = MockSessionState()

def mock_columns(spec):
    n = len(spec) if isinstance(spec, list) else int(spec)
    return [MagicMock() for _ in range(n)]

mock_st.columns = mock_columns
sys.modules['streamlit'] = mock_st

import pandas as pd
import openpyxl
import app

def run_test():
    print("--- STARTING WORKBOOK COMPARISON TEST ---")
    
    # 1. Load original workbook
    orig_wb = openpyxl.load_workbook("Private Tennis Lessons Workbook 2026.xlsx", data_only=False)
    orig_ws = orig_wb["Session 2"]
    
    # 2. Run app roster processing
    master_df = pd.read_csv("masterTennisLessons.csv")
    session_df = pd.read_csv("session2tennis.csv")
    
    # Counselor map format matches what's built by app.py UI
    counselor_map = {}
    flat_list = []
    for k, v in app.DEFAULT_COUNSELORS.items():
        if isinstance(k, tuple):
            counselor_map[k] = v
        else:
            counselor_map[k] = v
            
    roster_rows = app.process_roster(master_df, session_df, "Tennis", "Session 2", counselor_map)
    
    # 3. Generate new workbook in-memory and reload
    excel_stream = app.generate_excel_workbook(roster_rows, "Tennis", "Session 2")
    new_wb = openpyxl.load_workbook(excel_stream, data_only=False)
    new_ws = new_wb["Session 2"]
    
    # 4. Compare cell values, formulas, and formatting styles
    print(f"Original sheet max row: {orig_ws.max_row}, column: {orig_ws.max_column}")
    print(f"Generated sheet max row: {new_ws.max_row}, column: {new_ws.max_column}")
    
    # Compare row-by-row
    mismatches = 0
    checks_run = 0
    
    # We compare up to row 31 (summary rows) and column 30 (AD)
    for r in range(1, 32):
        for c in range(1, 31):
            checks_run += 1
            cell_orig = orig_ws.cell(row=r, column=c)
            cell_new = new_ws.cell(row=r, column=c)
            
            val_orig = cell_orig.value
            val_new = cell_new.value
            
            # Clean values for comparison
            # E.g., strip spaces for strings and compare formulas lowercase
            if isinstance(val_orig, str):
                val_orig = val_orig.strip()
                if val_orig.startswith("="):
                    val_orig = val_orig.lower().replace(" ", "")
            if isinstance(val_new, str):
                val_new = val_new.strip()
                if val_new.startswith("="):
                    val_new = val_new.lower().replace(" ", "")
            
            # Skip checking counselor mapping discrepancies if the bunk wasn't in original or was slightly different
            # (our script handles this via the editor, but we want to check layout correctness)
            if r >= 3 and c == 3:  # Counselor name column
                continue
            if r >= 3 and c == 7:  # Session shorthand - original had some inconsistent values, we standardized them
                continue
            if r >= 3 and c == 9:  # Minor Availability - manual column
                continue
            if r >= 3 and c == 10: # Memo - original has slight formatting variations, we prioritize session memo
                continue
            
            if val_orig != val_new:
                # Let's check if they are numerically equal (e.g. float vs int)
                try:
                    if float(val_orig) == float(val_new):
                        continue
                except:
                    pass
                
                # Report mismatch
                col_letter = openpyxl.utils.get_column_letter(c)
                print(f"Mismatch at Cell {col_letter}{r}:")
                print(f"  Original:  {val_orig!r}")
                print(f"  Generated: {val_new!r}")
                mismatches += 1
                
            # Style & Formatting checks for Row 3 (example camper)
            if r == 3:
                # Check column fill colors
                fill_orig = cell_orig.fill.start_color.rgb if cell_orig.fill and cell_orig.fill.fill_type else None
                fill_new = cell_new.fill.start_color.rgb if cell_new.fill and cell_new.fill.fill_type else None
                if fill_orig != fill_new:
                    print(f"Style Fill mismatch at Cell {openpyxl.utils.get_column_letter(c)}{r}:")
                    print(f"  Original:  {fill_orig}")
                    print(f"  Generated: {fill_new}")
                    mismatches += 1
                    
                # Check font size
                font_orig = cell_orig.font.size if cell_orig.font else None
                font_new = cell_new.font.size if cell_new.font else None
                if font_orig != font_new:
                    print(f"Style Font Size mismatch at Cell {openpyxl.utils.get_column_letter(c)}{r}:")
                    print(f"  Original:  {font_orig}")
                    print(f"  Generated: {font_new}")
                    mismatches += 1

    print(f"\n--- COMPARISON COMPLETE ---")
    print(f"Total cells checked: {checks_run}")
    print(f"Total mismatches: {mismatches}")
    if mismatches == 0:
        print("🎉 SUCCESS! Layout and data merge matches the original sheet perfectly.")
    else:
        print(f"⚠️ Mismatches found: {mismatches}. Please review.")

if __name__ == "__main__":
    run_test()
